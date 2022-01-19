import requests as r
from openpyxl import Workbook
from openpyxl import load_workbook

select_file = input("Select excel file:")
wb = load_workbook(select_file)
select_sheet = input("Select your sheet:")
list1 = wb[select_sheet]
# print(wb.sheetnames)

active_sheet = wb.active
a = list1['A1'].value

b = list1.cell(1,2).value = 2
print(b,type(b))

response = r.get(a)
assert response.status_code == 200
select_row = input("Select a row for read URLS:")
select_response_row = input("Select response row:")
if select_row == "A1" and select_response_row == "A2":
    for i in list1.rows:
        urls = r.get(i[0].value)
        print(urls,":", i[0].value)
        write = i[1].value = str(urls)
        wb.save(select_file)
elif select_row == "A2" and select_response_row == "A3":
    for i in list1.rows:
        urls = r.get(i[1].value)
        print(urls,":", i[1].value)
        write = i[2].value = str(urls)
        wb.save(select_file)
print("Complete")