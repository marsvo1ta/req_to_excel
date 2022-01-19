import requests as r
from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook("/Users/anton/Downloads/Blog links-Stfalcon.xlsx")
list1 = wb["Лист1"]
print(wb.sheetnames)

active_sheet = wb.active
a = list1['A1'].value

b = list1.cell(1,2).value = 2
print(b,type(b))

response = r.get(a)
assert response.status_code == 200
for i in list1.rows:
    urls = r.get(i[0].value)
    print(urls,":", i[0].value)
    write = i[1].value = str(urls)
    wb.save("/Users/anton/Downloads/Blog links-Stfalcon.xlsx")




